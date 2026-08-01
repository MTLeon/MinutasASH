from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from src.catalog_models import ClientRecord, ContactRecord, OrganizationRecord, ProjectCatalogRecord
from src.database import AppDatabase, normalize_name


CATALOG_HEADERS: dict[str, list[str]] = {
    "contacts": [
        "name", "initials", "email", "role", "organization", "phone", "active", "notes"
    ],
    "clients": [
        "legal_name", "short_name", "tax_id", "address", "primary_contact_name",
        "primary_contact_email", "primary_contact_phone", "active", "notes"
    ],
    "organizations": [
        "legal_name", "short_name", "tax_id", "address", "email", "phone", "active", "notes"
    ],
    "projects": [
        "code", "description", "client", "project_manager", "approver",
        "default_minute_taker", "default_location", "document_type", "discipline",
        "folder_path", "active"
    ],
}


@dataclass
class ImportIssue:
    row_number: int
    message: str
    values: dict[str, Any] = field(default_factory=dict)


@dataclass
class ImportSummary:
    catalog: str
    total_rows: int = 0
    imported: int = 0
    skipped: int = 0
    duplicates: int = 0
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.issues


def _bool_value(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().casefold() in {"1", "true", "sí", "si", "yes", "activo", "active"}


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depende del instalador.
        raise RuntimeError("La importación Excel requiere openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]
    finally:
        # En Windows, los libros abiertos en modo read_only mantienen el ZIP
        # interno bloqueado hasta cerrar explícitamente el Workbook.
        workbook.close()


def read_rows(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return _read_csv(source)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(source)
    raise ValueError("Solo se admiten archivos CSV o XLSX.")


def _write_csv(path: Path, headers: list[str], rows: Iterable[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in headers})
    return path


def _write_xlsx(path: Path, headers: list[str], rows: Iterable[dict[str, Any]]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("La exportación Excel requiere openpyxl.") from exc
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Catálogo"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(key, "") for key in headers])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        for index, header in enumerate(headers, start=1):
            width = max(len(header) + 2, 12)
            values = [str(sheet.cell(row=row, column=index).value or "") for row in range(2, min(sheet.max_row, 100) + 1)]
            if values:
                width = min(max(width, max(len(value) for value in values) + 2), 42)
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
        if sheet.max_row >= 2 and sheet.max_column >= 1:
            table = Table(displayName="CatalogoASH", ref=f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        sheet.freeze_panes = "A2"
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path
    finally:
        # Libera de forma determinista los recursos ZIP de openpyxl.
        workbook.close()



def create_import_template(catalog: str, destination: str | Path) -> Path:
    """Crea un archivo vacío con los encabezados admitidos por el importador."""
    if catalog not in CATALOG_HEADERS:
        raise ValueError("Catálogo no compatible.")
    target = Path(destination)
    if target.suffix.lower() == ".csv":
        return _write_csv(target, CATALOG_HEADERS[catalog], [])
    return _write_xlsx(target, CATALOG_HEADERS[catalog], [])

def export_catalog(database: AppDatabase, catalog: str, destination: str | Path) -> Path:
    if catalog not in CATALOG_HEADERS:
        raise ValueError("Catálogo no compatible.")
    rows: list[dict[str, Any]]
    if catalog == "contacts":
        rows = database.list_contact_records(include_inactive=True)
    elif catalog == "clients":
        rows = database.list_clients(include_inactive=True)
    elif catalog == "organizations":
        rows = database.list_organizations(include_inactive=True)
    else:
        rows = database.list_projects()
    target = Path(destination)
    if target.suffix.lower() == ".csv":
        result = _write_csv(target, CATALOG_HEADERS[catalog], rows)
    else:
        result = _write_xlsx(target, CATALOG_HEADERS[catalog], rows)
    database.log_audit("export", catalog, str(result), None, {"rows": len(rows)})
    return result



def _record_exists(database: AppDatabase, catalog: str, values: dict[str, Any]) -> bool:
    if catalog == "contacts":
        key = normalize_name(str(values.get("name") or ""))
        return bool(key) and any(normalize_name(row.get("name") or "") == key for row in database.list_contact_records(include_inactive=True))
    if catalog == "clients":
        key = normalize_name(str(values.get("legal_name") or ""))
        return bool(key) and any(normalize_name(row.get("legal_name") or "") == key for row in database.list_clients(include_inactive=True))
    if catalog == "organizations":
        key = normalize_name(str(values.get("legal_name") or ""))
        return bool(key) and any(normalize_name(row.get("legal_name") or "") == key for row in database.list_organizations(include_inactive=True))
    if catalog == "projects":
        code = str(values.get("code") or "").strip().upper()
        return bool(code) and database.get_project(code) is not None
    return False

def _import_one(database: AppDatabase, catalog: str, values: dict[str, Any]) -> None:
    if catalog == "contacts":
        record = ContactRecord(
            name=str(values.get("name") or ""),
            initials=values.get("initials") or None,
            email=values.get("email") or None,
            role=values.get("role") or None,
            organization=values.get("organization") or None,
            phone=values.get("phone") or None,
            active=_bool_value(values.get("active")),
            notes=values.get("notes") or None,
        )
        database.upsert_contact_record(record)
    elif catalog == "clients":
        database.upsert_client(
            ClientRecord(
                legal_name=str(values.get("legal_name") or ""),
                short_name=values.get("short_name") or None,
                tax_id=values.get("tax_id") or None,
                address=values.get("address") or None,
                primary_contact_name=values.get("primary_contact_name") or None,
                primary_contact_email=values.get("primary_contact_email") or None,
                primary_contact_phone=values.get("primary_contact_phone") or None,
                active=_bool_value(values.get("active")),
                notes=values.get("notes") or None,
            )
        )
    elif catalog == "organizations":
        database.upsert_organization(
            OrganizationRecord(
                legal_name=str(values.get("legal_name") or ""),
                short_name=values.get("short_name") or None,
                tax_id=values.get("tax_id") or None,
                address=values.get("address") or None,
                email=values.get("email") or None,
                phone=values.get("phone") or None,
                active=_bool_value(values.get("active")),
                notes=values.get("notes") or None,
            )
        )
    elif catalog == "projects":
        record = ProjectCatalogRecord(
            code=str(values.get("code") or ""),
            description=values.get("description") or None,
            client=values.get("client") or None,
            project_manager=values.get("project_manager") or None,
            approver=values.get("approver") or None,
            default_minute_taker=values.get("default_minute_taker") or None,
            default_location=values.get("default_location") or "Microsoft Teams",
            document_type=values.get("document_type") or "MRE",
            discipline=values.get("discipline") or "PR",
            folder_path=values.get("folder_path") or None,
            active=_bool_value(values.get("active")),
        )
        database.upsert_project_profile(record.model_dump())
    else:
        raise ValueError("Catálogo no compatible.")


def import_catalog(
    database: AppDatabase,
    catalog: str,
    source_path: str | Path,
    *,
    duplicate_policy: str = "upsert",
) -> ImportSummary:
    if catalog not in CATALOG_HEADERS:
        raise ValueError("Catálogo no compatible.")
    if duplicate_policy not in {"upsert", "skip"}:
        raise ValueError("La política de duplicados debe ser upsert o skip.")
    rows = read_rows(source_path)
    summary = ImportSummary(catalog=catalog, total_rows=len(rows))
    for row_number, row in enumerate(rows, start=2):
        try:
            if duplicate_policy == "skip" and _record_exists(database, catalog, row):
                summary.duplicates += 1
                summary.skipped += 1
                continue
            _import_one(database, catalog, row)
            summary.imported += 1
        except Exception as exc:
            summary.skipped += 1
            summary.issues.append(ImportIssue(row_number=row_number, message=str(exc), values=row))
    database.log_audit(
        "import",
        catalog,
        str(source_path),
        None,
        {"total": summary.total_rows, "imported": summary.imported, "skipped": summary.skipped, "duplicates": summary.duplicates, "duplicate_policy": duplicate_policy},
    )
    return summary
