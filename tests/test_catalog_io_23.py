from __future__ import annotations

import csv
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from src.catalog_io import create_import_template, export_catalog, import_catalog
from src.catalog_models import ContactRecord
from src.database import AppDatabase


class CatalogIO23Tests(unittest.TestCase):
    def test_csv_import_and_export(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            db = AppDatabase(root / "minutas.db")
            source = root / "contacts.csv"
            with source.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(
                    handle, fieldnames=["name", "email", "organization", "active"]
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "name": "Ana Pérez",
                        "email": "ana@example.com",
                        "organization": "ASH",
                        "active": "Sí",
                    }
                )
                writer.writerow(
                    {
                        "name": "",
                        "email": "sin-nombre@example.com",
                        "organization": "ASH",
                        "active": "Sí",
                    }
                )
            summary = import_catalog(db, "contacts", source)
            self.assertEqual(summary.imported, 1)
            self.assertEqual(summary.skipped, 1)
            destination = root / "export.csv"
            export_catalog(db, "contacts", destination)
            self.assertTrue(destination.is_file())
            self.assertIn("Ana Pérez", destination.read_text(encoding="utf-8-sig"))

    def test_creates_blank_xlsx_import_template(self) -> None:
        with TemporaryDirectory() as temp:
            destination = Path(temp) / "plantilla_clientes.xlsx"
            result = create_import_template("clients", destination)
            self.assertEqual(result, destination)
            self.assertTrue(destination.is_file())
            from openpyxl import load_workbook

            workbook = load_workbook(destination, read_only=True)
            try:
                headers = [cell.value for cell in next(workbook.active.iter_rows())]
                self.assertIn("legal_name", headers)
                self.assertIn("primary_contact_email", headers)
            finally:
                workbook.close()

            # Regresión Windows: el XLSX debe quedar liberado al terminar.
            renamed = destination.with_name("plantilla_clientes_liberada.xlsx")
            destination.replace(renamed)
            renamed.replace(destination)

    def test_xlsx_roundtrip_and_duplicate_skip(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source_db = AppDatabase(root / "source.db")
            source_db.upsert_contact_record(
                ContactRecord(name="Carlos Soto", email="carlos@example.com", organization="ASH")
            )
            workbook = root / "contacts.xlsx"
            export_catalog(source_db, "contacts", workbook)
            self.assertTrue(workbook.is_file())

            target_db = AppDatabase(root / "target.db")
            first = import_catalog(target_db, "contacts", workbook, duplicate_policy="skip")
            second = import_catalog(target_db, "contacts", workbook, duplicate_policy="skip")
            self.assertEqual(first.imported, 1)
            self.assertEqual(second.imported, 0)
            self.assertEqual(second.duplicates, 1)
            self.assertEqual(len(target_db.list_contact_records(include_inactive=True)), 1)

            # Regresión Windows: la lectura en modo read_only no puede dejar
            # abierto el ZIP interno de openpyxl.
            renamed = workbook.with_name("contacts_liberado.xlsx")
            workbook.replace(renamed)
            renamed.replace(workbook)

    def test_imports_split_name_user_export(self) -> None:
        from openpyxl import Workbook

        with TemporaryDirectory() as temp:
            source = Path(temp) / "usuarios.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "USR_ID",
                    "Nick",
                    "Rut",
                    "eMail",
                    "Primer N.",
                    "Segundo N.",
                    "Apellido Pat.",
                    "Apellido Mat.",
                ]
            )
            sheet.append(
                [225, "ADP", "11.111.111-1", "ana@example.test", "Ana", "Demo", "Pérez", "Prueba"]
            )
            sheet.append([173, "BDR", "22.222.222-2", "x", "Bruno", "Demo", "Rojas", "Prueba"])
            workbook.save(source)
            workbook.close()
            database = AppDatabase(Path(temp) / "catalog.db")
            summary = import_catalog(database, "contacts", source)
            rows = database.list_contact_records(include_inactive=True)
            self.assertEqual(summary.imported, 2)
            self.assertEqual(
                {row["name"] for row in rows},
                {"Ana Demo Pérez Prueba", "Bruno Demo Rojas Prueba"},
            )
            bruno = next(row for row in rows if row["name"].startswith("Bruno"))
            self.assertIsNone(bruno["email"])
            self.assertEqual(bruno["organization"], "ASH")
            self.assertIn("RUT:", bruno["notes"])


if __name__ == "__main__":
    unittest.main()
